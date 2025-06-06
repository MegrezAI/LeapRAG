#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

from openpyxl import load_workbook, Workbook
import sys
from io import BytesIO

from rag.nlp import find_codec

import pandas as pd


class RAGFlowExcelParser:
    def html(self, fnm, chunk_rows=256):

        # if isinstance(fnm, str):
        #     wb = load_workbook(fnm)
        # else:
        #     wb = load_workbook(BytesIO(fnm))++

        s_fnm = fnm
        if not isinstance(fnm, str):
            s_fnm = BytesIO(fnm)
        else:
            pass

        try:
            wb = load_workbook(s_fnm)
        except Exception as e:
            print(f'****wxy: file parser error: {e}, s_fnm={s_fnm}, trying convert files')
            df = pd.read_excel(s_fnm)
            wb = Workbook()
            # if len(wb.worksheets) > 0:
            #   del wb.worksheets[0]
            # else: pass
            ws = wb.active
            ws.title = "Data"
            for col_num, column_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=column_name)
            else:
                pass
            for row_num, row in enumerate(df.values, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                else:
                    pass
            else:
                pass

        tb_chunks = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue

            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{t.value}</th>"
            tb_rows_0 += "</tr>"

            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(
                  rows[1 + chunk_i * chunk_rows: 1 + (chunk_i + 1) * chunk_rows]
                ):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{c.value}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    async def __call__(self, fnm):
        # if isinstance(fnm, str):
        #   wb = load_workbook(fnm)
        # else:
        #   wb = load_workbook(BytesIO(fnm))

        s_fnm = fnm
        if not isinstance(fnm, str):
            s_fnm = BytesIO(fnm)
        else:
            pass

        try:
            wb = load_workbook(s_fnm)
        except Exception as e:
            print(f'****wxy: file parser error: {e}, s_fnm={s_fnm}, trying convert files')
            df = pd.read_excel(s_fnm)
            wb = Workbook()
            if len(wb.worksheets) > 0:
                del wb.worksheets[0]
            else:
                pass
            ws = wb.active
            ws.title = "Data"
            for col_num, column_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=column_name)
            else:
                pass
            for row_num, row in enumerate(df.values, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                else:
                    pass
            else:
                pass

        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            ti = list(rows[0])
            for r in list(rows[1:]):
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    fields.append(t)
                line = "; ".join(fields)
                if sheetname.lower().find("sheet") < 0:
                    line += " ——" + sheetname
                res.append(line)
        return res

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = load_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += len(list(ws.rows))
            return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    psr(sys.argv[1])

